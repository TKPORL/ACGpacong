# -*- coding: utf-8 -*-
"""
ACG游戏姬工具 - 抓取帖子标题、PC/AZ 分类、详情页移动云盘下载链接
默认只保留 PC / AZ 分类,过滤"喵笔记"。
支持:按日期筛选、指定页数、txt / csv / xlsx 输出。
"""

import argparse
import csv
import os
import random
import re
import sys
import time
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import requests
from bs4 import BeautifulSoup

BASE_URL = "https://acgyxj.top"
FALLBACK_URLS = [
    "https://acgyxj.cc",
    "https://acgyxj.xyz",
    "https://acgus.top",
    "https://acgyx.us",
]
DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Referer": "https://acgyx.us/",
}

# 运行时实际使用的 base_url，由 resolve_base_url() 设定
_resolved_base_url: Optional[str] = None


import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def _test_url(url: str, timeout: int = 10) -> bool:
    """快速测试一个 URL 是否可达"""
    try:
        resp = requests.get(url, timeout=timeout, headers=DEFAULT_HEADERS, allow_redirects=True, verify=False)
        return resp.status_code < 400
    except Exception:
        return False


def resolve_base_url() -> str:
    """探测可用域名：先试主站，不行再试备用域名"""
    global _resolved_base_url
    if _resolved_base_url:
        return _resolved_base_url
    # 测试主站
    if _test_url(BASE_URL):
        _resolved_base_url = BASE_URL
        return BASE_URL
    print(f"[备用] 主站 {BASE_URL} 不可达，尝试备用域名 ...", file=sys.stderr)
    for url in FALLBACK_URLS:
        if _test_url(url):
            _resolved_base_url = url
            DEFAULT_HEADERS["Referer"] = url + "/"
            print(f"[备用] 使用备用域名: {url}", file=sys.stderr)
            return url
    # 全部失败，返回主站（让后续请求自行报错）
    _resolved_base_url = BASE_URL
    print(f"[备用] 所有域名均不可达，使用主站 {BASE_URL}", file=sys.stderr)
    return BASE_URL


def get_base_url() -> str:
    """获取当前可用的 base_url"""
    return resolve_base_url()

# 需要排除的分类(教程类)
EXCLUDED_CATEGORIES = {"喵笔记"}


# --------------------- 工具函数 ---------------------

def http_get(url: str, session: requests.Session, retries: int = 3,
             timeout: int = 20) -> Optional[str]:
    for i in range(retries):
        try:
            resp = session.get(url, timeout=timeout, headers=DEFAULT_HEADERS, verify=False)
            resp.raise_for_status()
            resp.encoding = resp.apparent_encoding or "utf-8"
            return resp.text
        except Exception as e:
            print(f"  [重试 {i + 1}/{retries}] {url} 失败: {e}", file=sys.stderr)
            time.sleep(2 + i)
    return None


def parse_date(s: str) -> Optional[datetime]:
    """支持 YYYY.M.D / YYYY-M-D / YYYY/M/D / YYYY-MM-DD 等格式"""
    if not s:
        return None
    s = s.strip()
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d", "%Y.%-m.%-d", "%Y.%-m.%-d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def in_date_range(pub_time: str,
                  date_from: Optional[datetime],
                  date_to: Optional[datetime]) -> bool:
    """判断发布时间是否在指定区间内"""
    if not (date_from or date_to):
        return True
    # pub_time 可能是 "2026年6月16日" / "2026-06-16" / "2026.06.16"
    dt: Optional[datetime] = None
    m = re.search(r"(\d{4})[-./年](\d{1,2})[-./月](\d{1,2})", pub_time or "")
    if m:
        try:
            dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            dt = None
    if not dt:
        return True  # 解析不出日期时默认放行(避免漏抓)
    if date_from and dt < date_from:
        return False
    if date_to and dt > date_to:
        return False
    return True


# --------------------- 抓取核心 ---------------------

def fetch_list_page(page: int, session: requests.Session) -> List[Dict]:
    base = get_base_url()
    url = base + "/" if page == 1 else f"{base}/page/{page}"
    html = http_get(url, session)
    if not html:
        return []
    soup = BeautifulSoup(html, "lxml")
    posts: List[Dict] = []

    for article in soup.select("article"):
        title_a = article.select_one("h2 a, h3 a, .entry-title a, .post-title a")
        if not title_a:
            continue
        title = title_a.get_text(strip=True)
        link = (title_a.get("href") or "").strip()
        if not link:
            continue

        # 分类(PC / AZ / 喵笔记等)
        cat_a = article.select_one(".category .tags a, .cat, .entry-cat a")
        category = cat_a.get_text(strip=True) if cat_a else ""

        # 发布时间:位于 .list-post-author 中,例如 "@姬姬 2026年6月16日"
        pub_time = ""
        author_el = article.select_one(".list-post-author")
        if author_el:
            # 去掉前面的 @用户名,只保留日期
            import re as _re
            t = author_el.get_text(" ", strip=True)
            m = _re.search(r"(\d{4})[-./年](\d{1,2})[-./月](\d{1,2})", t)
            if m:
                pub_time = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        # 兜底:尝试其他常见选择器
        if not pub_time:
            for sel in ["time", ".date", ".entry-date", ".post-date", ".meta-date",
                        ".published", "[datetime]"]:
                t_el = article.select_one(sel)
                if t_el:
                    pub_time = (t_el.get("datetime") or t_el.get("title")
                                or t_el.get_text(strip=True))
                    if pub_time:
                        break

        posts.append({
            "title": title,
            "category": category,
            "url": link,
            "pub_time": pub_time,
        })
    return posts


def get_total_pages(session: requests.Session, max_check: int = 50) -> int:
    """探测总页数:从首页找最大页码链接"""
    html = http_get(get_base_url() + "/", session)
    if not html:
        return 1
    pages = [int(p) for p in re.findall(r"/page/(\d+)", html) if p.isdigit()]
    if pages:
        return max(pages)
    # 没有分页时查找 1 2 3 4 形式
    nums = [int(n) for n in re.findall(r">(\d+)</a>", html) if n.isdigit()]
    return max(nums) if nums else 1


def parse_links(html: str) -> Dict[str, object]:
    """从详情页解析下载链接、作弊码、汉化包解压码等
    返回:
      yun_links: List[str]   所有移动云盘链接(分卷时多)
      baidu_links: List[str] 所有百度网盘链接
      cheat_code: str        作弊码(只匹配"作弊码"这一关键字)
      unzip_code: str        汉化包解压码(只匹配"汉化包解压码"这一关键字)
    """
    soup = BeautifulSoup(html, "lxml")
    content = (soup.select_one(".single-content")
               or soup.select_one(".entry-content")
               or soup.select_one("article .content")
               or soup.select_one("article")
               or soup)
    text = content.get_text("\n", strip=True)

    result: Dict[str, object] = {
        "yun_links": [],
        "baidu_links": [],
        "baidu_codes": [],
        "cheat_code": "",
        "unzip_code": "",
    }

    # 移动云盘(可能多个,分卷)
    for m in re.finditer(r"https?://yun\.139\.com[^\s\)\"']*", text):
        link = m.group(0).rstrip("。,;；")
        if link not in result["yun_links"]:
            result["yun_links"].append(link)
    # 百度网盘(可能多个)
    for m in re.finditer(r"https?://pan\.baidu\.com/s/[^\s\)\"']*", text):
        link = m.group(0).rstrip("。,;；")
        if link not in result["baidu_links"]:
            result["baidu_links"].append(link)
    # 百度网盘提取码(可能多个,对应多个链接)
    for m in re.finditer(r"提取码[：: ]*([a-zA-Z0-9]{2,8})", text):
        code = m.group(1)
        if code not in result["baidu_codes"]:
            result["baidu_codes"].append(code)

    # 作弊码:仅匹配"作弊码"关键字(用户明确只要这一个)
    cc_m = re.search(r"作弊码[：: ]*([A-Za-z0-9]{2,16})", text)
    if cc_m:
        result["cheat_code"] = cc_m.group(1)

    # 汉化包解压码:仅匹配"汉化包解压码"关键字
    uz_m = re.search(r"汉化包解压码[：: ]*([A-Za-z0-9]{2,16})", text)
    if uz_m:
        result["unzip_code"] = uz_m.group(1)

    return result


# 后缀关键字(长->短优先匹配)
SUFFIX_KEYWORDS = [
    "内嵌AI汉化版", "内嵌汉化版", "内嵌版", "内嵌AI汉化",
    "AI汉化版", "AI汉化补丁", "AI汉化",
    "汉化版", "汉化补丁", "汉化补丁版",
    "官中步兵版", "官中赞助版", "官中版",
    "赞助版", "步兵版",
    "作弊码", "存档",
    "画廊MOD",
]


def _split_name_suffix(text: str) -> Tuple[str, str]:
    """把去掉末尾方括号后的文本切成 (游戏名, 后缀)
    边界:第一个匹配到的后缀关键字
    - 关键字前是游戏名,关键字开始算后缀
    - 找不到关键字,游戏名为空,后缀是整段
    """
    if not text:
        return "", ""
    for kw in sorted(SUFFIX_KEYWORDS, key=len, reverse=True):
        idx = text.find(kw)
        if idx > 0:  # 必须非开头(避免错误切分)
            return text[:idx].strip(), text[idx:].strip()
    return "", text.strip()


def _extract_chinese_name(text: str) -> str:
    """从游戏名段提取游戏名
    规则:
    - 有中文:取前 2 个含中文字符的段(跳过日文假名段)
    - 没中文:保留英文游戏名,但去掉版本号段(Ver/Release/v/WIP 开头的段)
    - 兜底:如果提取结果为空,返回原始文本
    """
    if not text:
        return ""
    parts = re.split(r"\s+", text.strip())

    # 检查整段是否有中文
    has_chinese = bool(re.search(r"[\u4e00-\u9fff]", text))

    if has_chinese:
        # 有中文:提取含中文的段
        out: List[str] = []
        for p in parts:
            if not p:
                continue
            if re.search(r"[\u3040-\u309f\u30a0-\u30ff]", p):
                continue
            if re.search(r"[\u4e00-\u9fff]", p):
                out.append(p)
                if len(out) >= 2:
                    break
        return " ".join(out) if out else text.strip()
    else:
        # 没中文:保留英文游戏名,去掉版本号段
        out: List[str] = []
        for p in parts:
            if not p:
                continue
            # 跳过版本号段(Ver/Release/v/WIP/Ep. 开头)
            if re.match(r"^(Ver|Release|v|WIP|Ep\.)", p, re.IGNORECASE):
                continue
            out.append(p)
        return " ".join(out) if out else text.strip()


def simplify_title(title: str, category: str,
                   cheat_code: str = "", unzip_code: str = "") -> str:
    """自动简化标题:
    1) 先识别并保留"末尾的大小标记"([X.XXG] / 【PC X.XXG】 等)
    2) 删【...】 / [...] 分类标签(剩余的)
    3) 删中英文括号内内容
    4) 删版本号 vX.Y.Z
    5) 删前缀 更新/新作/增添(可选+AZ)
    6) 符号替换 : / ? ? → 空格
    7) 找后缀关键字 → 切分 游戏名-后缀
    8) 游戏名取中文(无中文用整段)
    9) 末尾标准化:PC→[PC 大小]  /  AZ 保持原样
    10) 拼作弊码(原标题有"作弊码"字样时)
    11) 拼汉化包解压码(标题含"汉化"时)
    """
    t = (title or "").strip()

    # 0) 先识别"末尾的大小标记" — 中英文方括号都算
    # 区分:含"分类关键字(SLG/RPG/ACT/...)"的是分类标签,纯数字单位/MG是大小
    CLASS_WORDS = ("SLG|RPG|ACT|AVG|RTS|TPS|FPS|MOBA|探索|冒险|动作|色情|动漫|后宫|动态|"
                   "异种|奇幻|异世界|美少女|情色|幻想|射击|沙盒|家园|战略|休闲|卖春|援交|"
                   "魅魔|淫魔|堕落|塔防|解谜|策略|惊悚|恐怖|校园|恋爱|科幻|魔幻|玄幻|武侠|"
                   "历史|现代|未来|末日|都市|乡村|野外|室内|真人|3D|2D|像素|全动态|步兵|"
                   "赞助|官中|汉化|剧情|H|经营|养成|卡牌|消除|益智|音乐|舞蹈|竞速|体育|"
                   "格斗|搏击|拳击|摔角|摔跤|相扑|神作|新作|更新")
    size_match = None
    size_brackets = ""
    # 找标题末尾的方括号(中英都试)
    for pat in [r"\[([^\[\]]*)\]\s*$",
                "\u3010([^\u3010\u3011]*)\u3011" + r"\s*$"]:
        m = re.search(pat, t)
        if not m:
            continue
        content = m.group(1)
        is_class = re.search(CLASS_WORDS, content)
        is_size = (re.search(r"\d+\.?\d*\s*[MGK]", content, re.I)
                   or "PC+" in content
                   or content.strip() in {"PC", "安卓"})
        # 大小标记:含数字单位 且 不含类目词
        if is_size and not is_class:
            size_match = m
            size_brackets = m.group(0)
            t = t[:m.start()].strip()
            break

    # 1) 删【...】 / [...] 分类标签(剩余的)
    t = re.sub("\u3010[^\u3010\u3011]*\u3011", "", t)
    t = re.sub(r"\[[^\[\]]*\]", "", t)
    # 2) 删中英文括号内内容
    t = re.sub(r"[（(][^)）]*[)）]", "", t)
    # 3) 删版本号 (v1.2.3, v0.9.7, v0.2.6.0c 等)
    t = re.sub(r"\s*v\d+[\w.]*\b", " ", t, flags=re.IGNORECASE)
    # 4) 删前缀 更新/新作/增添(可带 AZ)
    t = re.sub(r"^(更新|新作|增添)(?:\s*[A-Z]+)?\s*", "", t)
    # 5) 符号替换:全位置 : / ? ? , & → 空格
    for sym in [":", "/", "?", "?", ",", "&"]:
        t = t.replace(sym, " ")
    # 合并空白
    t = re.sub(r"\s+", " ", t).strip()

    # 7) 切分 游戏名 / 后缀
    name_raw, suffix = _split_name_suffix(t)

    # 8) 游戏名取中文
    if name_raw:
        name = _extract_chinese_name(name_raw)
    else:
        # 没切出后缀关键字 → 整段交给 _extract_chinese_name 智能提取
        # 有中文则提中文名，无中文则保留整段
        name = _extract_chinese_name(t)
        suffix = ""

    # 9) 末尾标准化
    if category == "PC" and size_match:
        # PC 分类:统一为 [PC <大小>]
        m_size = re.search(r"(\d+\.?\d*\s*[MGK])", size_match.group(1), re.I)
        if m_size:
            end_bracket = f"[PC {m_size.group(1).strip()}]"
        else:
            end_bracket = f"[PC {size_match.group(1).strip()}]"
    else:
        # AZ / 无大小:保留方括号形态,但括号里的 : / ? ? , & 仍要变空格
        end_bracket = size_brackets
        for sym in [":", "/", "?", "?", ",", "&"]:
            end_bracket = end_bracket.replace(sym, " ")

    # 10) 拼作弊码:原标题必须含"作弊码"字样
    if cheat_code and "作弊码" in (title or ""):
        if f"作弊码{cheat_code}" in suffix:
            pass  # 作弊码已在标题中，无需重复添加
        elif "作弊码" in suffix:
            suffix = suffix.replace("作弊码", f"作弊码{cheat_code}", 1)
        else:
            suffix = (suffix + "作弊码" + cheat_code) if suffix else ("作弊码" + cheat_code)

    # 11) 拼汉化包解压码:标题含"汉化"
    extras: List[str] = []
    if unzip_code and "汉化" in (title or ""):
        extras.append(f"汉化包解压码{unzip_code}")

    # 12) 拼装
    main_parts: List[str] = []
    if name and suffix:
        # 复合标签(含 -)用空格连接避免 -- 歧义,简单标签用 - 连接
        sep = " " if "-" in suffix else "-"
        main_parts.append(f"{name.rstrip('-')}{sep}{suffix.rstrip()}")
    elif name:
        main_parts.append(name)
    elif suffix:
        main_parts.append(suffix.rstrip())
    if extras:
        main_parts.append("".join(extras))

    # 主段 + 末尾方括号:中文【】紧贴,英[ ] 前留 1 空格
    main = " ".join(p for p in main_parts if p).strip()
    if end_bracket and main:
        sep = "" if end_bracket.startswith("\u3010") else " "
        return f"{main}{sep}{end_bracket}".strip()
    if end_bracket:
        return end_bracket.strip()
    return main if main else (title or "").strip()


def format_title(title: str, category: str,
                 cheat_code: str = "", unzip_code: str = "") -> str:
    """简化标题(包装函数,保持向后兼容)"""
    return simplify_title(title, category, cheat_code, unzip_code)


def fetch_post_detail(post: Dict, session: requests.Session) -> Dict:
    """抓取详情页:补充下载链接 + 处理标题(简化/PC 标记/作弊码/汉化码)"""
    html = http_get(post["url"], session)
    if not html:
        return {**post, "yun_links": [], "baidu_links": [],
                "baidu_codes": [], "cheat_code": "", "unzip_code": "",
                "title": post["title"]}
    info = parse_links(html)
    new_title = simplify_title(
        post["title"], post.get("category", ""),
        str(info.get("cheat_code", "") or ""),
        str(info.get("unzip_code", "") or ""),
    )
    return {**post, **info, "title": new_title}


# --------------------- 收集与过滤 ---------------------

def collect_posts(pages: int, session: requests.Session, delay: float,
                  exclude_cats: set) -> List[Dict]:
    all_posts: List[Dict] = []
    for p in range(1, pages + 1):
        print(f"[列表] 抓取第 {p}/{pages} 页 ...", file=sys.stderr)
        posts = fetch_list_page(p, session)
        if not posts:
            print(f"  第 {p} 页为空,停止翻页", file=sys.stderr)
            break
        all_posts.extend(posts)
        time.sleep(delay + random.random() * 0.3)
    print(f"[列表] 共 {len(all_posts)} 个帖子", file=sys.stderr)
    return all_posts


def filter_posts(posts: List[Dict], exclude_cats: set,
                 date_from: Optional[datetime],
                 date_to: Optional[datetime]) -> List[Dict]:
    out: List[Dict] = []
    for p in posts:
        if p["category"] in exclude_cats:
            continue
        if not in_date_range(p["pub_time"], date_from, date_to):
            continue
        out.append(p)
    return out


# --------------------- 输出 ---------------------

def build_txt_filename(date_from: Optional[datetime],
                       date_to: Optional[datetime]) -> str:
    """按抓取的时间段生成 txt 文件名
    单日: acgyx_20260616.txt
    范围: acgyx_20260610_20260616.txt
    无日期: acgyx_<时间戳>.txt
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if date_from and date_to and date_from == date_to:
        return f"acgyx_{date_from.strftime('%Y%m%d')}.txt"
    if date_from and date_to:
        return (f"acgyx_{date_from.strftime('%Y%m%d')}"
                f"_{date_to.strftime('%Y%m%d')}.txt")
    if date_from:
        return f"acgyx_{date_from.strftime('%Y%m%d')}.txt"
    return f"acgyx_{ts}.txt"


def save_txt(rows: List[Dict], path: str):
    """txt 格式:每条记录多行(标题/分类/发布时间/下载链接N/原文链接)"""
    with open(path, "w", encoding="utf-8") as f:
        # 文件头(用户要求)
        f.write("Tsinho爬取工具\n")
        if not rows:
            f.write("(无匹配结果)\n")
            return
        for i, r in enumerate(rows):
            if i > 0:
                f.write("\n")
            f.write(f"标题:{r.get('title', '')}\n")
            f.write(f"分类:{r.get('category', '')}\n")
            f.write(f"发布时间:{r.get('pub_time', '')}\n")
            yun_links = r.get("yun_links") or []
            for idx, link in enumerate(yun_links, 1):
                f.write(f"下载链接{idx}:{link}\n")
            if not yun_links:
                f.write("下载链接:(无)\n")
            f.write(f"原文链接:{r.get('url', '')}\n")


def save_csv(rows: List[Dict], path: str):
    fields = ["title", "category", "pub_time", "url",
              "yun_links", "baidu_links", "baidu_codes", "cheat_code"]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            row = dict(r)
            # 列表字段用 "|" 分隔
            for k in ("yun_links", "baidu_links", "baidu_codes"):
                v = row.get(k)
                if isinstance(v, list):
                    row[k] = "|".join(v)
                elif v is None:
                    row[k] = ""
            w.writerow(row)


def save_excel(rows: List[Dict], path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("[警告] openpyxl 未安装,跳过 Excel", file=sys.stderr)
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "ACG游戏姬"
    headers_cn = ["帖子标题", "分类", "发布时间", "帖子链接",
                  "移动云盘链接", "百度网盘链接", "作弊码"]
    ws.append(headers_cn)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center")
    widths = [70, 8, 16, 50, 60, 60, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for r in rows:
        yun = "\n".join(r.get("yun_links") or []) or ""
        bd = "\n".join(r.get("baidu_links") or []) or ""
        ws.append([
            r.get("title", ""), r.get("category", ""),
            r.get("pub_time", ""), r.get("url", ""),
            yun, bd, r.get("cheat_code", ""),
        ])
    for row in ws.iter_rows(min_row=2):
        for idx in (3, 4, 5):
            cell = row[idx]
            if cell.value and "\n" not in str(cell.value):
                cell.hyperlink = cell.value
                cell.font = Font(color="0000FF", underline="single")
    wb.save(path)


def render_txt(rows: List[Dict]) -> str:
    """返回 txt 文本(用于网页/API 实时返回)"""
    buf: List[str] = ["Tsinho爬取工具"]
    if not rows:
        buf.append("(无匹配结果)")
        return "\n".join(buf) + "\n"
    for i, r in enumerate(rows):
        if i > 0:
            buf.append("")
        buf.append(f"标题:{r.get('title', '')}")
        buf.append(f"分类:{r.get('category', '')}")
        buf.append(f"发布时间:{r.get('pub_time', '')}")
        yun_links = r.get("yun_links") or []
        if yun_links:
            for idx, link in enumerate(yun_links, 1):
                buf.append(f"下载链接{idx}:{link}")
        else:
            buf.append("下载链接:(无)")
        buf.append(f"原文链接:{r.get('url', '')}")
    return "\n".join(buf) + "\n"


# --------------------- CLI 入口 ---------------------

def main():
    ap = argparse.ArgumentParser(description="ACG游戏姬帖子抓取工具")
    ap.add_argument("-p", "--pages", type=int, default=0,
                    help="要抓取的页数(0 表示自动探测到末页)")
    ap.add_argument("--max-pages", type=int, default=20,
                    help="自动模式下的最大页数上限(防失控)")
    ap.add_argument("-o", "--output", default="acgyx_posts",
                    help="输出文件名前缀(默认 acgyx_posts)")
    ap.add_argument("-d", "--delay", type=float, default=1.0,
                    help="两次请求之间的基础延迟秒数(默认 1.0)")
    ap.add_argument("--date", type=str, default=None,
                    help="筛选某一天的帖子(支持 2026.6.16 / 2026-06-16)")
    ap.add_argument("--date-from", type=str, default=None,
                    help="起始日期(含)")
    ap.add_argument("--date-to", type=str, default=None,
                    help="结束日期(含)")
    ap.add_argument("--limit", type=int, default=0,
                    help="最多抓取详情页的数量(0 表示不限制)")
    ap.add_argument("--format", choices=["txt", "csv", "xlsx", "all"],
                    default="txt", help="输出格式(默认 txt)")
    ap.add_argument("--include-miaobiji", action="store_true",
                    help="包含'喵笔记'分类(默认排除)")
    ap.add_argument("--no-detail", action="store_true",
                    help="不抓详情页,只输出列表信息")
    args = ap.parse_args()

    exclude = set() if args.include_miaobiji else EXCLUDED_CATEGORIES
    date_from = parse_date(args.date) or parse_date(args.date_from)
    date_to = parse_date(args.date) or parse_date(args.date_to)
    if date_from and not date_from and args.date_from:
        print(f"[警告] 无法解析起始日期: {args.date_from}", file=sys.stderr)

    session = requests.Session()
    session.headers.update(DEFAULT_HEADERS)

    # 决定页数
    if args.pages <= 0:
        total = get_total_pages(session)
        pages = min(total, args.max_pages)
        print(f"[分页] 自动探测到 {total} 页,将抓取前 {pages} 页",
              file=sys.stderr)
    else:
        pages = args.pages

    posts = collect_posts(pages, session, args.delay, exclude)
    posts = filter_posts(posts, exclude, date_from, date_to)
    print(f"[筛选] 符合条件的帖子: {len(posts)} 条", file=sys.stderr)

    if args.no_detail:
        rows = posts
    else:
        limit = args.limit if args.limit > 0 else len(posts)
        total_n = min(len(posts), limit)
        print(f"[详情] 0/{total_n} ...", file=sys.stderr)
        rows = []
        for i, p in enumerate(posts[:limit], 1):
            print(f"  [{i}/{total_n}] {p['title'][:40]}", file=sys.stderr)
            rows.append(fetch_post_detail(p, session))
            if i % 10 == 0:
                print(f"[详情] 进度 {i}/{total_n}", file=sys.stderr)
            time.sleep(args.delay + random.random() * 0.3)
        print(f"[详情] 完成 {len(rows)} 条", file=sys.stderr)

    # 输出
    fmts = ["txt", "csv", "xlsx"] if args.format == "all" else [args.format]
    if "txt" in fmts:
        # 文件名按日期段生成(用户要求)
        if args.output == "acgyx_posts":
            path = build_txt_filename(date_from, date_to)
        else:
            path = args.output + ".txt"
        save_txt(rows, path)
        print(f"[保存] TXT -> {path}", file=sys.stderr)
    if "csv" in fmts:
        path = args.output + ".csv"
        save_csv(rows, path)
        print(f"[保存] CSV -> {path}", file=sys.stderr)
    if "xlsx" in fmts:
        path = args.output + ".xlsx"
        save_excel(rows, path)
        print(f"[保存] Excel -> {path}", file=sys.stderr)

    yun = sum(1 for r in rows if r.get("yun_links"))
    bd = sum(1 for r in rows if r.get("baidu_links"))
    print(f"[统计] 总数: {len(rows)} | 含移动云盘: {yun} | 含百度网盘: {bd}",
          file=sys.stderr)


if __name__ == "__main__":
    main()
